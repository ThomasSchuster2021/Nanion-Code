import sys
import os
from os import path
import string

import numpy as np
import openpyxl
import math
from scipy import stats
from scipy.optimize import curve_fit
from numpy import linspace, random, log, exp, sqrt, var
import matplotlib.pyplot as plt
import scipy.io


# gets columns from sheet relevent for cell_num specified
def read_pharm(sht, cell_num):
    col_names = [[i, sht.cell(row=1, column=i).value] for i in range(1, sht.max_column)]
    col_names = [s for s in col_names if s[1] != None]
    specific_cols = [c[0] for c in col_names if 'Sweep' in c[1] or 'Time' in c[1] or '(%d)' % cell_num in c[1]]
    data = []
    for c in specific_cols:
        data.append([sht.cell(row=i, column=c).value for i in range(1, sht.max_row + 1)])
    dataset = {}
    for d in data:
        key = d[0].replace('(%d)' % cell_num, '')
        dataset[key] = d[1:]
    return dataset


# writes data in a column to excel sheet starting at cell specific by row and column numbers
def write_to_col(col, row, data, sht):
    for i in range(len(data)):
        sht.cell(row=row + i, column=col).value = data[i]


# writes data in a row to excel sheet starting at cell specific by row and column numbers
def write_to_row(col, row, data, sht):
    for i in range(len(data)):
        sht.cell(row=row, column=col + i).value = data[i]


# converts xls to xlsx
def xls_to_xlsx(infile):
    outfile = infile + 'x'
    file = open(infile, 'r')
    fid = file.readlines()
    data = [f.split('\t') for f in fid]

    wb = openpyxl.Workbook()
    sht = wb['Sheet']
    for i, d in enumerate(data):
        write_to_row(1, i + 1, d, sht)
    wb.save(outfile)
    wb.close()
    return outfile


# returns dictionary of traces with keys in format Trace_1_<datatype>_<sweepnum>_<cellnum>
def get_traces(infile):
    m = scipy.io.loadmat(infile) #loads the matlab file called infile
    trace_names = []
    for x in m.keys():
        if 'Trace' in x:
            trace_names.append(x)
    traces = {}
    for n in trace_names:
        traces[n] = m[n]
    return traces

#def doseresponse0(filename, cell_num):
def doseresponse0(run, cell_num):
    data_type = 3
    filename = run + '.mat'
    t_num = filename.split('_')[0]
    r_num = filename.split('.')[0]
    traces = get_traces(filename)

    #excel_filename = '%s/%s_1NaPharmUDB.xls' % (os.getcwd(), r_num)
    #excel_filename = 'C:\\Users\\thoma\\OneDrive\\Documents\\Research Code\\Data\\' + run + '_1NaPharmUDB.xls'
    excel_filename = run + '_1NaPharmUDB.xls'
    xlsx_name = xls_to_xlsx(excel_filename)
    wb = openpyxl.load_workbook(xlsx_name)
    sht_name = wb.sheetnames[0]
    sht = wb[sht_name]
    pharm_data = read_pharm(sht, cell_num)
    sweep_vals = []
    for a, s, c in zip(pharm_data['Age'], pharm_data['Sweep'], pharm_data['Concentration']):
        sweep_vals.append(['Trace_' + s + '_%d' % (cell_num), int(a), c])

    # find transition points
    t_2pts = []
    t_pt = []
    ct = 0
    l = len(sweep_vals)
    for i, s in enumerate(sweep_vals):
        if i > 2 and s[1] < sweep_vals[i - 1][1]:
            prev_s = sweep_vals[i - 1]
            pt = [prev_s]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 2], prev_s]
            t_pt.append(pt)
            t_2pts.append(pts)
        if i == l - 1:
            pt = sweep_vals[i]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 1], pt]
            t_pt.append(pt)
            t_2pts.append(pts)
        ct = ct + 1

    # plot trace for each dose
    j = 0
    latecurrent0 = []
    peakcurrent0 = []
    peak_locs0 = []
    counter = 0
    # to use 2 traces for each dose
    for p in t_2pts:
        sumlate0 = 0
        sumpeak0 = 0
        print(counter)
        for x in p:

            ysinrange = []
            xsinrange_time = []
            xsinrange_points = []
            xs0 = []
            ys0 = []
            trace_num = x[0]
            if trace_num in traces:
                trace = traces[trace_num]
            else:
                ct = ct - 1
                trace_num = 'Trace_' + str(1) + '_' + str(data_type) + '_' + str(ct) + '_' + str(cell_num)
                trace = traces[trace_num]
            xs = [x[0] for x in trace]
            ys = [x[1] for x in trace]

            # 0 Hz peak calculation
            k = 0

            for y in ys:
                #print('y: ')
                #print(type(y))
                if k >= 0:
                    if k < 2500:  # 2500 data points per spike (10 Hz = 0.1 sec/spike) (25000 points/sec * 0.1 sec/spike = 2500 points/spike)
                        xsinrange_time.append(xs[k])
                        xsinrange_points.append(k)
                        ysinrange.append(y)
                k = k + 1
            peak = min(ysinrange)
            #print('Peak: ')
            #print(peak)
            sumpeak0 = sumpeak0 + peak


            # find the x location of the peak
            i = 0
            if j < 4:
                for x1 in xsinrange_points:
                    if ysinrange[i] == peak:
                        peak_loc = x1
                        peak_locs0.append(peak_loc)
                    i = i + 1
            if j > 3:
                np.array(peak_locs0)
                peak_locu = sum(peak_locs0) / len(peak_locs0)
                peak_loc = int(peak_locu)

            # specify which points to include for late current and compute late current
            z = 0
            start_loc = peak_loc + 350
            end_loc = peak_loc + 450
            for y1 in ysinrange:
                if z > start_loc:
                    if z < end_loc:
                        sumlate0 = sumlate0 + ysinrange[z]
                z = z + 1
        #print('Sum of Peaks: ')
        #print(sumpeak0)
        #print(' ')
        counter = counter + 1
        sumlate0 = sumlate0 / 2
        sumpeak0 = sumpeak0 / 2
        latecurrent0.append(sumlate0)
        peakcurrent0.append(sumpeak0)
        j = j + 1
    latecurrent0f = np.array(latecurrent0)
    peakcurrent0f = np.array(peakcurrent0)
    print(peakcurrent0f)
    peakcurrent0f = peakcurrent0f / min(peakcurrent0f) #min because most negative
    peakcurrent0f = peakcurrent0f * 100
    latecurrent0f = latecurrent0f * 100
    return([peakcurrent0f, latecurrent0f])

def doseresponse10(run, cell_num): #the second value is much more accurate
    data_type = 3
    filename = run + ".mat"
    t_num = filename.split('_')[0]
    r_num = filename.split('.')[0]
    traces = get_traces(filename)

    excel_filename = run + '_1NaPharmUDB.xls'
    xlsx_name = xls_to_xlsx(excel_filename)
    wb = openpyxl.load_workbook(xlsx_name)
    sht_name = wb.sheetnames[0]
    sht = wb[sht_name]
    pharm_data = read_pharm(sht, cell_num)
    sweep_vals = []
    for a, s, c in zip(pharm_data['Age'], pharm_data['Sweep'], pharm_data['Concentration']):
        sweep_vals.append(['Trace_' + s + '_%d' % (cell_num), int(a), c])

    # find transition points
    t_2pts = []
    t_pt = []
    ct = 0
    l = len(sweep_vals)
    for i, s in enumerate(sweep_vals):
        if i > 2 and s[1] < sweep_vals[i - 1][1]:
            prev_s = sweep_vals[i - 1]
            pt = [prev_s]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 2], prev_s]
            t_pt.append(pt)
            t_2pts.append(pts)
        if i == l - 1:
            pt = sweep_vals[i]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 1], pt]
            t_pt.append(pt)
            t_2pts.append(pts)
        ct = ct + 1

    # plot trace for each dose
    j = 0
    latecurrent10 = []
    peakcurrent10 = []
    peak_locs10 = []

    # to use 2 traces for each dose
    for p in t_2pts:
        sumlate10 = 0
        sumpeak10 = 0
        for x in p:
            ysinrange = []
            xsinrange_time = []
            xsinrange_points = []
            xs10 = []
            ys10 = []
            trace_num = x[0]
            if trace_num in traces:
                trace = traces[trace_num]
            else:
                ct = ct - 1
                trace_num = 'Trace_' + str(1) + '_' + str(data_type) + '_' + str(ct) + '_' + str(cell_num)
                trace = traces[trace_num]
            xs = [x[0] for x in trace]
            ys = [x[1] for x in trace]

            # 10 Hz peak calculation
            k = 0
            for y in ys:
                if k >= 125000:
                    if k <= 127500:  # 2500 data points per spike (10 Hz = 0.1 sec/spike) (25000 points/sec * 0.1 sec/spike = 2500 points/spike)
                        xsinrange_time.append(xs[k])
                        xsinrange_points.append(k)
                        ysinrange.append(y)
                k = k + 1
            peak = min(ysinrange)
            #print('Peak: ')
            #print(peak)
            sumpeak10 = sumpeak10 + peak

            # find the x location of the peak
            i = 0
            if j < 4:
                for x1 in xsinrange_points:
                    if ysinrange[i] == peak:
                        peak_loc = i
                        peak_locs10.append(peak_loc)
                    i = i + 1
            if j > 3:
                np.array(peak_locs10)
                peak_locu = sum(peak_locs10) / len(peak_locs10)
                peak_loc = int(peak_locu)

            # specify which points to include for late current and compute late current
            z = 125000
            start_loc = peak_loc + 350
            end_loc = peak_loc + 450
            for y in ysinrange:
                if z > start_loc:
                    if z < end_loc:
                        sumlate10 = sumlate10 + y
                z = z + 1
        #print('Sum of Peaks: ')
        #print(sumpeak10)
        #print(' ')
        sumlate10 = sumlate10 / 2
        sumpeak10 = sumpeak10 / 2
        latecurrent10.append(sumlate10)
        peakcurrent10.append(sumpeak10)
        j = j + 1
    latecurrent10f = np.array(latecurrent10)
    peakcurrent10f = np.array(peakcurrent10)
    peakcurrent10f = peakcurrent10f / min(peakcurrent10f)
    peakcurrent10f = peakcurrent10f * 100
    latecurrent10f = latecurrent10f * 100
    return ([peakcurrent10f, latecurrent10f])

def doseresponse20(run, cell_num):
    filename = run + ".mat"
    data_type = 3
    t_num = filename.split('_')[0]
    r_num = filename.split('.')[0]
    traces = get_traces(filename)

    excel_filename = run + '_1NaPharmUDB.xls'
    xlsx_name = xls_to_xlsx(excel_filename)
    wb = openpyxl.load_workbook(xlsx_name)
    sht_name = wb.sheetnames[0]
    sht = wb[sht_name]
    pharm_data = read_pharm(sht, cell_num)
    sweep_vals = []
    for a, s, c in zip(pharm_data['Age'], pharm_data['Sweep'], pharm_data['Concentration']):
        sweep_vals.append(['Trace_' + s + '_%d' % (cell_num), int(a), c])

    # find transition points
    t_2pts = []
    t_pt = []
    ct = 0
    l = len(sweep_vals)
    for i, s in enumerate(sweep_vals):
        if i > 2 and s[1] < sweep_vals[i - 1][1]:
            prev_s = sweep_vals[i - 1]
            pt = [prev_s]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 2], prev_s]
            t_pt.append(pt)
            t_2pts.append(pts)
        if i == l - 1:
            pt = sweep_vals[i]
            pts = [prev_s] if i == 1 else [sweep_vals[i - 1], pt]
            t_pt.append(pt)
            t_2pts.append(pts)
        ct = ct + 1

    # plot trace for each dose
    j = 0
    latecurrent20 = []
    peakcurrent20 = []
    peak_locs20 = []

    # to use 2 traces for each dose
    for p in t_2pts:
        sumlate20 = 0
        sumpeak20 = 0
        for x in p:
            ysinrange = []
            xsinrange_time = []
            xsinrange_points = []
            xs20 = []
            ys20 = []
            trace_num = x[0]
            if trace_num in traces:
                trace = traces[trace_num]
            else:
                ct = ct - 1
                trace_num = 'Trace_' + str(1) + '_' + str(data_type) + '_' + str(ct) + '_' + str(cell_num)
                trace = traces[trace_num]
            xs = [x[0] for x in trace]
            ys = [x[1] for x in trace]

            # 20 Hz peak calculation
            k = 0
            for y in ys:
                if k >= 195000: # why is this number so high?
                    if k < 196250:  # 2500 data points per spike (20 Hz = 0.05 sec/spike) (25000 points/sec * 0.05 sec/spike = 1250 points/spike)
                    #if k < 197500:
                        xsinrange_time.append(xs[k])
                        xsinrange_points.append(k)
                        ysinrange.append(y)
                k = k + 1
            peak = min(ysinrange)
            print('Peak: ')
            print(peak)
            sumpeak20 = sumpeak20 + peak

            # find the x location of the peak
            i = 0
            if j < 4:
                for x1 in xsinrange_points:
                    if ysinrange[i] == peak:
                        peak_loc = i
                        peak_locs20.append(peak_loc)
                    i = i + 1
            if j > 3:
                np.array(peak_locs20)
                peak_locu = sum(peak_locs20) / len(peak_locs20)
                peak_loc = int(peak_locu)

            # specify which points to include for late current and compute late current
            z = 195000
            start_loc = peak_loc + 350
            end_loc = peak_loc + 450
            for y1 in ysinrange:
                if z > start_loc:
                    if z < end_loc:
                        sumlate20 = sumlate20 + y1
                z = z + 1
        print('Sum of Peaks: ')
        print(sumpeak20)
        print(' ')
        sumlate20 = sumlate20 / 2
        sumpeak20 = sumpeak20 / 2
        latecurrent20.append(sumlate20)
        peakcurrent20.append(sumpeak20)
        j = j + 1
    latecurrent20f = np.array(latecurrent20)
    peakcurrent20f = np.array(peakcurrent20)
    peakcurrent20f = peakcurrent20f / min(peakcurrent20f)
    peakcurrent20f = peakcurrent20f * 100
    latecurrent20f = latecurrent20f * 100
    return ([peakcurrent20f, latecurrent20f])
